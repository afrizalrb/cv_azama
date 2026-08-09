# -*- coding: utf-8 -*-
"""
Migrasi satu kali dari Database CV. Azama Sejahtera (.xlsx) ke CSV siap impor
ke Google Sheets.

Jalankan:
    py scripts/migrate_excel.py
    py scripts/migrate_excel.py --input data/Database_CV_Azama_Sejahtera.xlsx --outdir data/csv

Skrip ini TIDAK menyentuh Google sama sekali. Output berupa satu file CSV per
tab, plus ringkasan baris masuk / dilewati beserta alasannya.

Catatan penting soal apa yang SENGAJA tidak dimigrasi:

1. stock_movements hanya berisi baris `opening` dari STOK_PRODUK.
   Kolom "Stok Akhir" di Excel sudah memperhitungkan seluruh penjualan historis.
   Kalau 111 invoice lama ikut ditulis sebagai `sale_out`, stok jadi terhitung
   dua kali dan langsung minus.

2. gallon_ledger dibiarkan kosong (hanya header).
   Galon dari transaksi 2024-2025 secara fisik sudah ditukar berkali-kali.
   Membangkitkan `gallon_out` untuk seluruh riwayat akan memunculkan saldo
   3.000-an galon yang tidak pernah ada. Saldo galon dihitung bersih mulai
   tanggal sistem dipakai.
"""

import argparse
import csv
import hashlib
import os
import re
import secrets
import sys
import uuid
from collections import OrderedDict, defaultdict
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl belum terpasang. Jalankan: py -m pip install openpyxl")


# ---------------------------------------------------------------------------
# Konfigurasi posisi baris di file Excel sumber (sudah diverifikasi langsung)
# ---------------------------------------------------------------------------

BARIS_DATA = {
    "MASTER_PRODUK": 7,      # header di baris 6
    "MASTER_CUSTOMER": 8,    # header di baris 7
    "PENJUALAN": 13,         # header di baris 12
    "STOK_PRODUK": 9,        # header di baris 8
    "STOK_BAHAN": 9,         # header di baris 8
}

# Produk yang dipertahankan. Sisanya dibuang karena HPP kosong / tidak pernah dijual.
PRODUK_DIPERTAHANKAN = {"GD19", "GD15", "GD12", "GD05", "GTDS019"}

# GTDS019 tidak punya HPP di Excel. Keputusan: HPP = 70% harga jual (margin 30%).
MARGIN_GTDS019 = 0.30

# Kode "produk" yang sebenarnya hanya varian harga dari produk induk.
# Selisih harganya dipindah ke customer_prices, bukan jadi produk terpisah.
NORMALISASI_RESELLER = {
    "GD19R": "GD19",
    "GD19R2": "GD19",
    "GD12R": "GD12",
}

TEMPO_DEFAULT = 30

# Kode bahan dibuat manual karena STOK_BAHAN tidak punya kolom kode terisi.
KODE_BAHAN = {
    "galon 19 liter kosong": ("GLN19", "pcs"),
    "galon 12 liter kosong": ("GLN12", "pcs"),
    "galon 5 liter kosong": ("GLN05", "pcs"),
    "botol 600ml kosong": ("BTL600", "pcs"),
    "botol 120ml kosong": ("BTL120", "pcs"),
    "shrink plastic": ("SHRINK", "roll"),
}

# Seed user awal. sales_person_name harus persis sama dengan kolom Marketing
# di MASTER_CUSTOMER, karena itulah kunci pencocokan hak akses sales.
SEED_USERS = [
    ("admin", "Administrator", "admin", ""),
    ("zhulham", "Zhulham", "sales", "Zhulham"),
    ("abah", "Abah", "sales", "Abah"),
    ("mama", "Mama", "sales", "Mama"),
    ("produksi", "Staf Produksi", "produksi", ""),
]


# ---------------------------------------------------------------------------
# Pengumpul peringatan
# ---------------------------------------------------------------------------

class Laporan:
    """Menampung statistik dan peringatan supaya bisa dicetak rapi di akhir."""

    def __init__(self):
        self.masuk = OrderedDict()
        self.dilewati = defaultdict(list)
        self.peringatan = []
        self.tindakan = []

    def catat_masuk(self, tab, jumlah):
        self.masuk[tab] = jumlah

    def catat_lewat(self, tab, identitas, alasan):
        self.dilewati[tab].append((identitas, alasan))

    def warn(self, pesan):
        self.peringatan.append(pesan)

    def todo(self, pesan):
        self.tindakan.append(pesan)


LAP = Laporan()


# ---------------------------------------------------------------------------
# Utilitas
# ---------------------------------------------------------------------------

def bersih(nilai):
    """Ubah sel Excel jadi string rapi tanpa spasi berlebih."""
    if nilai is None:
        return ""
    return re.sub(r"\s+", " ", str(nilai)).strip()


def ke_angka(nilai, default=0):
    """Ambil angka dari sel yang bisa saja berisi 'Rp ', '-', atau teks kosong."""
    if nilai is None:
        return default
    if isinstance(nilai, (int, float)):
        return float(nilai)
    teks = re.sub(r"[^\d,.\-]", "", str(nilai)).replace(",", ".")
    if teks in ("", "-", ".", "-."):
        return default
    try:
        return float(teks)
    except ValueError:
        return default


def bulat(nilai):
    """Uang dan qty di sistem ini selalu bilangan bulat."""
    return int(round(ke_angka(nilai)))


def normalisasi_kode_produk(kode):
    """
    Samakan penulisan kode produk yang tidak konsisten di Excel.

    GD019 -> GD19   (nol di depan angka tiga digit dibuang)
    GD05  -> GD05   (tetap, karena memang produk 5 liter)
    GD19R2 -> GD19  (varian harga reseller, bukan produk berbeda)
    """
    k = bersih(kode).upper().replace(" ", "")
    if not k:
        return ""
    if k in NORMALISASI_RESELLER:
        return NORMALISASI_RESELLER[k]
    cocok = re.fullmatch(r"GD0(\d{2})", k)
    if cocok:
        return "GD" + cocok.group(1)
    return k


def volume_ke_ml(teks):
    """'19 L' / '15 liter' / '600 ml' -> jumlah mililiter."""
    t = bersih(teks).lower()
    angka = re.search(r"([\d.,]+)", t)
    if not angka:
        return 0
    n = ke_angka(angka.group(1))
    if "ml" in t:
        return int(round(n))
    return int(round(n * 1000))


def tempo_ke_hari(teks):
    """'30 hari' -> 30. Kosong -> default, dan dicatat sebagai peringatan."""
    n = ke_angka(teks, default=0)
    return int(n) if n > 0 else TEMPO_DEFAULT


def tanggal_str(nilai):
    """Sel tanggal Excel -> 'YYYY-MM-DD'. Mengembalikan None kalau bukan tanggal."""
    if isinstance(nilai, datetime):
        return nilai.strftime("%Y-%m-%d")
    if hasattr(nilai, "year"):
        return nilai.strftime("%Y-%m-%d")
    return None


def hash_password(password, salt):
    """
    Harus identik dengan Auth.gs di Apps Script:
    SHA-256 dari (password + salt), disimpan sebagai hex huruf kecil.
    """
    return hashlib.sha256((password + salt).encode("utf-8")).hexdigest()


def tulis_csv(outdir, nama, header, baris):
    path = os.path.join(outdir, nama + ".csv")
    # utf-8-sig supaya nama customer tetap benar saat diimpor lewat Google Sheets
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(baris)
    return path


# ---------------------------------------------------------------------------
# Pembacaan tiap tab
# ---------------------------------------------------------------------------

def baca_produk(wb):
    """MASTER_PRODUK -> dict kode -> data produk. Produk tanpa HPP dibuang."""
    ws = wb["MASTER_PRODUK"]
    produk = OrderedDict()

    for row in ws.iter_rows(min_row=BARIS_DATA["MASTER_PRODUK"], values_only=True):
        kode_asli = bersih(row[2])
        if not kode_asli:
            continue

        kode = normalisasi_kode_produk(kode_asli)
        nama = bersih(row[1])
        harga = bulat(row[6])

        if kode_asli.upper() in NORMALISASI_RESELLER:
            LAP.catat_lewat(
                "products", kode_asli,
                f"varian harga reseller -> dinormalisasi ke {kode}, "
                f"selisih harga dipindah ke customer_prices"
            )
            continue

        if kode not in PRODUK_DIPERTAHANKAN:
            LAP.catat_lewat(
                "products", kode_asli,
                "HPP kosong / tidak pernah muncul di transaksi"
            )
            continue

        if kode == "GTDS019":
            hpp = int(round(harga * (1 - MARGIN_GTDS019)))
            LAP.warn(
                f"GTDS019 tidak punya HPP di Excel. Diisi {hpp:,} "
                f"(margin {int(MARGIN_GTDS019*100)}% dari harga jual {harga:,}) "
                f"sesuai keputusan. Verifikasi angka ini ke bagian produksi."
            )
        else:
            hpp = bulat(row[5])

        produk[kode] = {
            "code": kode,
            "name": nama,
            "packaging_type": bersih(row[3]).title(),
            "volume_ml": volume_ke_ml(row[4]),
            "cogs": hpp,
            "price": harga,
            "min_stock": 0,
            "is_returnable": "TRUE",     # seluruh produk yang tersisa berkemasan galon
            "deposit_amount": 0,
            "is_active": "TRUE",
        }

    LAP.catat_masuk("products", len(produk))
    LAP.todo("Isi kolom deposit_amount di tab products — Excel tidak punya datanya.")
    LAP.todo("Isi kolom min_stock di tab products agar alert stok minimum berfungsi.")
    return produk


def baca_customer(wb):
    """MASTER_CUSTOMER -> dict kode -> data customer."""
    ws = wb["MASTER_CUSTOMER"]
    customers = OrderedDict()

    for row in ws.iter_rows(min_row=BARIS_DATA["MASTER_CUSTOMER"], values_only=True):
        kode = bersih(row[2]).upper()
        if not kode:
            continue

        tempo_asli = bersih(row[6])
        if not tempo_asli:
            LAP.warn(
                f"Customer {kode} ({bersih(row[1])}) tidak punya tempo pembayaran. "
                f"Diisi {TEMPO_DEFAULT} hari."
            )
        sales = bersih(row[8])
        if not sales:
            LAP.warn(
                f"Customer {kode} ({bersih(row[1])}) tidak punya sales. "
                f"Customer ini tidak akan terlihat oleh user role sales mana pun."
            )

        customers[kode] = {
            "code": kode,
            "name": bersih(row[1]),
            "area": bersih(row[3]),
            "type": bersih(row[4]),
            "payment_term_days": tempo_ke_hari(tempo_asli),
            "phone": bersih(row[7]),
            "sales_person": sales,
            "is_active": "TRUE",
            "_produk_khusus": normalisasi_kode_produk(row[5]) if bersih(row[5]) else "",
        }

    LAP.catat_masuk("customers", len(customers))
    return customers


def baca_penjualan(wb, produk, customers):
    """
    PENJUALAN -> sales_orders + sales_order_items.

    Satu baris Excel = satu produk. Baris dengan no invoice sama digabung
    menjadi satu order dengan beberapa item.
    """
    ws = wb["PENJUALAN"]
    orders = OrderedDict()
    total_baris = 0

    for row in ws.iter_rows(min_row=BARIS_DATA["PENJUALAN"], values_only=True):
        invoice = bersih(row[3]).upper()
        if not invoice:
            continue
        total_baris += 1

        tgl = tanggal_str(row[2])
        if not tgl:
            LAP.catat_lewat("sales_orders", invoice, "kolom Tanggal bukan tanggal valid")
            continue

        kode_cust = bersih(row[4]).upper()
        if kode_cust not in customers:
            LAP.catat_lewat(
                "sales_orders", invoice,
                f"customer '{kode_cust}' tidak ada di MASTER_CUSTOMER"
            )
            continue

        kode_asli = bersih(row[5]).upper()
        kode_prod = normalisasi_kode_produk(kode_asli)
        if kode_prod not in produk:
            LAP.catat_lewat(
                "sales_order_items", f"{invoice}/{kode_asli}",
                f"produk '{kode_asli}' tidak ada di daftar produk yang dipertahankan"
            )
            continue

        qty = bulat(row[6])
        harga = bulat(row[7])
        if qty <= 0 or harga <= 0:
            LAP.catat_lewat(
                "sales_order_items", f"{invoice}/{kode_asli}",
                f"qty ({qty}) atau harga ({harga}) tidak wajar"
            )
            continue

        if invoice not in orders:
            orders[invoice] = {
                "invoice_no": invoice,
                "order_date": tgl,
                "customer_code": kode_cust,
                "items": [],
            }
        elif orders[invoice]["customer_code"] != kode_cust:
            LAP.warn(
                f"Invoice {invoice} muncul dengan dua customer berbeda "
                f"({orders[invoice]['customer_code']} dan {kode_cust}). "
                f"Baris kedua diabaikan."
            )
            continue

        orders[invoice]["items"].append({
            "product_code": kode_prod,
            "product_code_asli": kode_asli,
            "qty": qty,
            "unit_price": harga,
            "unit_cogs": produk[kode_prod]["cogs"],
            "line_total": qty * harga,
        })

    orders = OrderedDict(
        sorted(orders.items(), key=lambda kv: (kv[1]["order_date"], kv[0]))
    )

    LAP.warn(
        f"Kolom Status di PENJUALAN kosong pada seluruh {total_baris} baris. "
        f"Semua invoice historis diberi status 'paid' dengan asumsi transaksi "
        f"lama sudah lunas. WAJIB diverifikasi manual — kalau ada yang sebenarnya "
        f"belum dibayar, aging piutang akan salah sejak hari pertama."
    )
    return orders, total_baris


def bangun_customer_prices(orders, produk, customers):
    """
    Harga khusus tidak ditulis eksplisit di Excel — kolom 'Produk Harga Khusus'
    hanya menyebut kode produknya, tanpa nominal.

    Jadi harga khusus diturunkan dari riwayat transaksi: kalau seorang customer
    konsisten membayar dengan harga berbeda dari harga master, selisih itulah
    harga khususnya. Yang dipakai adalah harga pada transaksi terakhir.
    """
    terakhir = {}
    for inv in orders.values():
        for item in inv["items"]:
            kunci = (inv["customer_code"], item["product_code"])
            sebelumnya = terakhir.get(kunci)
            if sebelumnya is None or inv["order_date"] >= sebelumnya[0]:
                terakhir[kunci] = (inv["order_date"], item["unit_price"])

    baris = []
    for (kode_cust, kode_prod), (tgl, harga) in sorted(terakhir.items()):
        harga_master = produk[kode_prod]["price"]
        if harga != harga_master:
            baris.append([kode_cust, kode_prod, harga])
            LAP.warn(
                f"Harga khusus terdeteksi dari riwayat: {kode_cust} "
                f"({customers[kode_cust]['name']}) membeli {kode_prod} "
                f"seharga {harga:,} (master {harga_master:,}), transaksi {tgl}."
            )
            # Selisih harga di atas dua kali lipat hampir pasti salah tulis kode
            # produk, bukan kesepakatan harga. Contoh nyata: GD05 (master 7.500)
            # tercatat 20.000, yang persis sama dengan harga GTDS005 Galon Mini.
            if harga_master and (harga / harga_master >= 2 or harga_master / harga >= 2):
                LAP.todo(
                    f"CURIGAI SALAH KODE PRODUK: {kode_cust} "
                    f"({customers[kode_cust]['name']}) tercatat membeli {kode_prod} "
                    f"seharga {harga:,} padahal harga masternya {harga_master:,}. "
                    f"Selisihnya terlalu jauh untuk sebuah harga khusus. Cek nota "
                    f"aslinya — kemungkinan yang dijual sebenarnya produk lain."
                )

    # Customer yang ditandai punya harga khusus di master tapi tidak terbukti di riwayat
    for kode, c in customers.items():
        prod_khusus = c["_produk_khusus"]
        if prod_khusus and not any(b[0] == kode and b[1] == prod_khusus for b in baris):
            LAP.todo(
                f"Customer {kode} ({c['name']}) ditandai punya harga khusus untuk "
                f"{prod_khusus} di Excel, tapi nominalnya tidak pernah ditulis dan "
                f"riwayatnya memakai harga normal. Isi manual di tab customer_prices "
                f"kalau memang ada kesepakatan harga khusus."
            )

    LAP.catat_masuk("customer_prices", len(baris))
    return baris


def baca_stok_produk(wb, produk):
    """STOK_PRODUK -> satu baris `opening` per produk di stock_movements."""
    ws = wb["STOK_PRODUK"]
    baris = []
    urut = 0
    saat_ini = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for row in ws.iter_rows(min_row=BARIS_DATA["STOK_PRODUK"], values_only=True):
        kode = normalisasi_kode_produk(row[2])
        if not kode:
            continue
        if kode not in produk:
            LAP.catat_lewat("stock_movements", kode, "produk tidak dipertahankan")
            continue

        qty = bulat(row[7])   # kolom "Stok Akhir"
        urut += 1
        baris.append([
            f"MOV{urut:05d}", saat_ini, "product", kode, qty, "opening",
            "migration", "", "Saldo awal hasil migrasi dari kolom Stok Akhir Excel",
            "migration", saat_ini,
        ])

    # Produk yang ada di master tapi tidak muncul di STOK_PRODUK tetap perlu
    # baris pembuka, supaya stok awalnya eksplisit nol, bukan tidak terdefinisi.
    sudah = {b[3] for b in baris}
    for kode in produk:
        if kode not in sudah:
            urut += 1
            baris.append([
                f"MOV{urut:05d}", saat_ini, "product", kode, 0, "opening",
                "migration", "", "Tidak tercatat di STOK_PRODUK, saldo awal diset 0",
                "migration", saat_ini,
            ])
            LAP.warn(f"Produk {kode} tidak ada di STOK_PRODUK. Stok awal diset 0.")

    LAP.catat_masuk("stock_movements", len(baris))
    return baris


def baca_bahan(wb):
    """STOK_BAHAN -> materials. Hanya nama item; angka di Excel masih kosong."""
    ws = wb["STOK_BAHAN"]
    baris = []
    urut = 0

    for row in ws.iter_rows(min_row=BARIS_DATA["STOK_BAHAN"], values_only=True):
        nama = bersih(row[2])
        if not nama:
            continue
        kunci = nama.lower()
        kode_manual = bersih(row[3]).upper()
        if kunci in KODE_BAHAN:
            kode, satuan = KODE_BAHAN[kunci]
        elif kode_manual:
            kode, satuan = kode_manual, bersih(row[5]) or "pcs"
        else:
            urut += 1
            kode, satuan = f"MAT{urut:03d}", bersih(row[5]) or "pcs"
            LAP.warn(f"Bahan '{nama}' tidak punya kode. Dibuatkan kode otomatis {kode}.")
        baris.append([kode, nama, satuan, bulat(row[11])])

    LAP.catat_masuk("materials", len(baris))
    LAP.todo(
        "STOK_BAHAN di Excel hanya berisi nama item — stok awal, satuan, dan "
        "stok minimal semuanya kosong. Lakukan opname bahan sekali, lalu input "
        "sebagai mutasi `opening` lewat menu Stok."
    )
    return baris


def bangun_users():
    """Seed user awal beserta password acak yang dicetak sekali ke layar."""
    baris = []
    kredensial = []
    for username, nama, role, sales_name in SEED_USERS:
        password = secrets.token_urlsafe(9)
        salt = secrets.token_hex(16)
        baris.append([
            str(uuid.uuid4()), username, hash_password(password, salt), salt,
            nama, role, sales_name, "TRUE",
        ])
        kredensial.append((username, password, role))
    LAP.catat_masuk("users", len(baris))
    return baris, kredensial


# ---------------------------------------------------------------------------
# Penulisan seluruh CSV
# ---------------------------------------------------------------------------

HEADER = {
    "users": ["user_id", "username", "password_hash", "salt", "full_name",
              "role", "sales_person_name", "is_active"],
    "products": ["code", "name", "packaging_type", "volume_ml", "cogs", "price",
                 "min_stock", "is_returnable", "deposit_amount", "is_active"],
    "customers": ["code", "name", "area", "type", "payment_term_days", "phone",
                  "sales_person", "is_active"],
    "customer_prices": ["customer_code", "product_code", "special_price"],
    "sales_orders": ["order_id", "invoice_no", "order_date", "customer_code",
                     "due_date", "status", "subtotal", "created_by", "created_at"],
    "sales_order_items": ["item_id", "order_id", "product_code", "qty",
                          "unit_price", "unit_cogs", "line_total"],
    "payments": ["payment_id", "order_id", "payment_date", "amount", "method",
                 "reference", "created_by"],
    "stock_movements": ["movement_id", "moved_at", "item_type", "item_code", "qty",
                        "movement_type", "ref_type", "ref_id", "notes",
                        "created_by", "created_at"],
    "gallon_ledger": ["ledger_id", "moved_at", "customer_code", "product_code",
                      "qty", "movement_type", "deposit_amount", "ref_type",
                      "ref_id", "notes", "created_by"],
    "production_batches": ["batch_id", "batch_no", "produced_at", "product_code",
                           "qty", "ph_value", "tds_value", "pic", "notes", "created_by"],
    "materials": ["code", "name", "unit", "min_stock"],
    "suppliers": ["code", "name", "address", "phone", "payment_term_days"],
    "material_purchases": ["purchase_id", "purchase_date", "supplier_code",
                           "material_code", "qty", "unit_price", "total", "created_by"],
    "operational_expenses": ["expense_id", "expense_date", "category", "amount",
                             "description", "created_by"],
    "audit_log": ["log_id", "timestamp", "user_id", "username", "action",
                  "payload", "result"],
}

TAB_KOSONG = ["payments", "gallon_ledger", "production_batches", "suppliers",
              "material_purchases", "operational_expenses", "audit_log"]


def main():
    ap = argparse.ArgumentParser(description="Migrasi Excel Azama ke CSV")
    ap.add_argument("--input", default="data/Database_CV_Azama_Sejahtera.xlsx")
    ap.add_argument("--outdir", default="data/csv")
    args = ap.parse_args()

    if not os.path.isfile(args.input):
        sys.exit(f"File sumber tidak ditemukan: {args.input}")
    os.makedirs(args.outdir, exist_ok=True)

    print(f"Membaca {args.input} ...")
    wb = openpyxl.load_workbook(args.input, data_only=True, read_only=True)

    produk = baca_produk(wb)
    customers = baca_customer(wb)
    orders, baris_penjualan = baca_penjualan(wb, produk, customers)
    harga_khusus = bangun_customer_prices(orders, produk, customers)
    mutasi_stok = baca_stok_produk(wb, produk)
    bahan = baca_bahan(wb)
    users, kredensial = bangun_users()

    # --- rakit sales_orders + sales_order_items ---
    saat_ini = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    baris_order, baris_item = [], []
    n_order = n_item = 0

    for inv in orders.values():
        if not inv["items"]:
            LAP.catat_lewat("sales_orders", inv["invoice_no"], "tidak punya item valid")
            continue
        n_order += 1
        order_id = f"ORD{n_order:05d}"
        tempo = customers[inv["customer_code"]]["payment_term_days"]
        jatuh_tempo = (
            datetime.strptime(inv["order_date"], "%Y-%m-%d") + timedelta(days=tempo)
        ).strftime("%Y-%m-%d")
        subtotal = sum(i["line_total"] for i in inv["items"])

        baris_order.append([
            order_id, inv["invoice_no"], inv["order_date"], inv["customer_code"],
            jatuh_tempo, "paid", subtotal, "migration", saat_ini,
        ])
        for item in inv["items"]:
            n_item += 1
            baris_item.append([
                f"ITM{n_item:05d}", order_id, item["product_code"], item["qty"],
                item["unit_price"], item["unit_cogs"], item["line_total"],
            ])

    LAP.catat_masuk("sales_orders", n_order)
    LAP.catat_masuk("sales_order_items", n_item)

    # --- tulis semua file ---
    data = {
        "users": users,
        "products": [[p[k] for k in HEADER["products"]] for p in produk.values()],
        "customers": [[c[k] for k in HEADER["customers"]] for c in customers.values()],
        "customer_prices": harga_khusus,
        "sales_orders": baris_order,
        "sales_order_items": baris_item,
        "stock_movements": mutasi_stok,
        "materials": bahan,
    }
    for tab in TAB_KOSONG:
        data[tab] = []

    print(f"\nMenulis CSV ke {args.outdir}/ ...")
    for tab in HEADER:
        tulis_csv(args.outdir, tab, HEADER[tab], data.get(tab, []))
        print(f"  {tab}.csv  ({len(data.get(tab, []))} baris data)")

    cetak_ringkasan(baris_penjualan, subtotal_total=sum(o[6] for o in baris_order))
    cetak_kredensial(args.outdir, kredensial)
    simpan_kredensial(args.outdir, kredensial)


def cetak_ringkasan(baris_penjualan, subtotal_total):
    print("\n" + "=" * 72)
    print("RINGKASAN MIGRASI")
    print("=" * 72)

    print("\nBaris masuk per tab:")
    for tab, n in LAP.masuk.items():
        print(f"  {tab:<22} {n:>5}")

    if LAP.dilewati:
        print("\nBaris dilewati:")
        for tab, daftar in LAP.dilewati.items():
            print(f"  [{tab}]")
            for identitas, alasan in daftar:
                print(f"    - {identitas}: {alasan}")

    print(f"\nTotal omzet historis terbawa: Rp {subtotal_total:,}")
    print(f"Baris PENJUALAN dibaca      : {baris_penjualan}")

    if LAP.peringatan:
        print("\n" + "-" * 72)
        print("PERINGATAN (baca sebelum data dipakai)")
        print("-" * 72)
        for i, p in enumerate(LAP.peringatan, 1):
            print(f"  {i}. {p}")

    if LAP.tindakan:
        print("\n" + "-" * 72)
        print("PERLU DIISI MANUAL SETELAH IMPOR")
        print("-" * 72)
        for i, t in enumerate(LAP.tindakan, 1):
            print(f"  {i}. {t}")


def cetak_kredensial(outdir, kredensial):
    print("\n" + "=" * 72)
    print("PASSWORD AWAL — DICETAK SEKALI SAJA, TIDAK BISA DIPULIHKAN")
    print("=" * 72)
    print("Yang tersimpan di users.csv hanya hash-nya. Catat sekarang.\n")
    for username, password, role in kredensial:
        print(f"  {username:<12} {password:<16} ({role})")
    print("\nGanti seluruh password ini lewat menu Users setelah sistem berjalan.")
    print(f"\nSeluruh CSV berada di {outdir}/ — folder data/ sudah masuk .gitignore,")
    print("jadi nama customer dan omzet tidak akan ikut ter-push ke repo publik.")


def simpan_kredensial(outdir, kredensial):
    """
    Simpan password awal ke berkas di folder data/ (sudah di-gitignore).

    Alasannya praktis: hash di users.csv tidak bisa dibalik, dan output terminal
    gampang tertutup. Kalau berkas ini hilang sebelum password sempat dicatat,
    satu-satunya jalan adalah menjalankan ulang migrasi user.
    """
    path = os.path.join(outdir, "KREDENSIAL_AWAL.txt")
    with open(path, "w", encoding="utf-8") as f:
        f.write("PASSWORD AWAL SISTEM AZAMA\n")
        f.write(f"Dibuat: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        for username, password, role in kredensial:
            f.write(f"{username:<12} {password:<16} ({role})\n")
        f.write(
            "\nBerkas ini berisi password polos. Hapus segera setelah seluruh\n"
            "pengguna mengganti passwordnya lewat menu Users.\n"
            "Berkas ini berada di folder data/ yang sudah masuk .gitignore,\n"
            "sehingga tidak akan ikut ter-push ke GitHub.\n"
        )
    print(f"Password juga disimpan di {path}")


if __name__ == "__main__":
    main()
